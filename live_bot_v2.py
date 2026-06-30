"""
╔══════════════════════════════════════════════════════════════╗
║  SPH/SPL Structural Pivot Bot v2.2                          ║
║  ─────────────────────────────────────────────────────────── ║
║  Strategy:  15-Min SPH/SPL Breakout (Always-In Reversal)    ║
║  Style:     Positional (hold overnight)                     ║
║  Orders:    BUY / SELL / SYNTHETIC / ZEBRA (configurable)   ║
║  Sizing:    Configurable lots per leg                       ║
║  Expiry:    Next-Week (auto-selected)                       ║
║  Pivots:    w=3 Symmetric (ILM-inspired, fewer whipsaws)    ║
╚══════════════════════════════════════════════════════════════╝
"""

import os
import sys
import datetime
import pandas as pd
import numpy as np
import time
import json
import logging
import threading
import urllib.parse
import urllib.request
import requests
from kiteconnect import KiteConnect, KiteTicker
from bot_secrets import get_kite_api_key, load_secrets, require_env
from kite_auth import KiteLoginError, ensure_access_token, load_kite_env, read_access_token, token_is_valid

# ═══════════════════════════════════════════════════
# CONFIGURATION — CHANGE THESE BEFORE RUNNING
# ═══════════════════════════════════════════════════

# --- MODE ---
BOT_MODE = 'SIMULATION'          # Options: 'SIMULATION' or 'LIVE'
                           # SIMULATION = logs signals only, no real orders
                           # LIVE       = places real market orders via Kite

# --- ORDER TYPE ---
ORDER_TYPE = 'BUY'         # Options: 'BUY', 'SELL', 'SYNTHETIC', 'ZEBRA'
                           #
                           # BUY       = Buy 1 ITM option (CE for LONG, PE for SHORT)
                           #             Risk: premium paid. Reward: unlimited.
                           #
                           # SELL      = Sell 1 OTM option (PE for LONG, CE for SHORT)
                           #             Risk: unlimited. Reward: premium collected.
                           #
                           # SYNTHETIC = Buy ITM + Sell OTM (same expiry)
                           #             LONG: Buy ITM CE + Sell OTM PE
                           #             SHORT: Buy ITM PE + Sell OTM CE
                           #             Risk: defined. Reward: defined.
                           #
                           # ZEBRA     = Buy 2 Deep ITM + Sell 1 ATM (Zero Extrinsic Back Ratio)
                           #             LONG: Buy 2x Deep ITM CE + Sell 1x ATM CE
                           #             SHORT: Buy 2x Deep ITM PE + Sell 1x ATM PE
                           #             Cost: near-zero extrinsic. Behaves like futures.

# --- POSITION SIZING ---
NUM_LOTS = 5              # Number of lots per leg (1 lot = 65 qty for NIFTY)
                           # For ZEBRA: Buy leg = 2x this, Sell leg = 1x this

# --- STRIKE SELECTION ---
ITM_OFFSET = 1             # How many strikes ITM for BUY leg (1 = ATM±50)
OTM_OFFSET = 1             # How many strikes OTM for SELL leg (1 = ATM±50)
ZEBRA_DEEP_ITM = 5         # Strikes deep ITM for ZEBRA buy leg (5 = ATM±250)

# --- STARTUP OVERRIDE ---
# Force a direction at boot when resuming from state.
# Options: None, 'LONG', 'SHORT'
STARTUP_DIRECTION_OVERRIDE = 'None'

# --- PROFIT BOOKING + STRUCTURAL RE-ENTRY ---
# Book the trade once option premium gain reaches PROFIT_BOOK_PCT (e.g. 0.60 => +60%).
# After booking, optionally re-enter SAME direction at fresh 1 ITM only if price is
# still beyond the active level (LONG: close > SPH, SHORT: close < SPL). Else stay flat.
PROFIT_BOOK_ENABLED = True
PROFIT_BOOK_PCT = 0.60                   # 60% of entry premium
REENTRY_REQUIRES_LEVEL_BREAK = True       # if False, always re-enter same direction
BOOK_REENTRY_CUTOFF = (15, 0)             # (hh, mm) — block fresh re-entries after this
BOOK_REENTRY_COOLDOWN_SEC = 5             # min seconds between successive book/re-entry checks

# --- EXECUTION TIMING ---
# True  = SPH/SPL entries/reversals only on each 15-min candle close (chart-aligned).
# False = tick-by-tick structural signals (immediate on SPH/SPL cross).
EXECUTE_ON_15MIN_CLOSE = False

# Min seconds between structural reversals (0 = immediate tick reversals).
REVERSAL_COOLDOWN_SEC = 0

# Cap trailing SPH/SPL to today's session high/low (old behavior). False keeps raw structure.
SESSION_SL_CAP_ENABLED = False

# --- WEBSOCKET SCHEDULE (IST, Mon–Fri) ---
TICKER_CONNECT_TIME = (8, 0)        # connect before market open
TICKER_DISCONNECT_TIME = (15, 30)   # disconnect after close; positions held overnight

# Telegram alert on each 15-min pivot refresh (raw vs active levels).
PIVOT_TELEGRAM_ALERTS = True

# Configure Logger
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("bot_v2_execution.log"),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger("LiveBotV2")


class StructuralPivotBot:
    """
    SPH/SPL Breakout Bot — Simplified Structural Pivot Strategy.

    Pivots:  Calculated on 15-min structure (SPH/SPL, window=4)
    Entry:   Tick-by-tick on SPH/SPL cross (set EXECUTE_ON_15MIN_CLOSE for bar-close only)
    Exit:    Reversal only (always-in system) or Expiry Day at 15:15
    SL:      Trailing — SPL for Longs, SPH for Shorts (updates as pivots move)
    """

    def __init__(self, mode='SIMULATION'):
        self.mode = mode
        self.lot_size = 65
        self.num_lots = NUM_LOTS
        self.qty = self.lot_size * self.num_lots
        self.order_type = ORDER_TYPE        # 'BUY', 'SELL', 'SYNTHETIC', 'ZEBRA'
        self.itm_offset = ITM_OFFSET        # strikes ITM for buy leg
        self.otm_offset = OTM_OFFSET        # strikes OTM for sell leg
        self.zebra_deep_itm = ZEBRA_DEEP_ITM

        # --- Kite Connection ---
        self.kite = None
        self.kws = None
        self.instrument_token = 256265  # NIFTY 50 (NSE Index)
        load_secrets()
        self.api_key = get_kite_api_key()
        self.access_token = None

        # --- Instruments ---
        self.instruments_df = pd.DataFrame()
        self.expiry_date = None

        # --- Position State ---
        self.positions = []   # Max 1 position at a time
        self.trades_log = []

        # --- Pivot State ---
        self.sph = np.nan
        self.spl = np.nan

        # --- Candle Manager ---
        self.current_minute = None
        self.current_candle = {'open': 0, 'high': 0, 'low': 0, 'close': 0}
        self.df_1min = pd.DataFrame()
        self.df_15min = pd.DataFrame()  # Native 15-min candles from Kite
        self.current_15min_key = None   # Track live 15-min candle boundary
        self.current_15min_candle = {'open': 0, 'high': 0, 'low': 0, 'close': 0}

        # --- Execution Control ---
        self.last_exec_15min = None   # Prevents double-fire for expiry exit
        self.last_scan_log = None     # Throttle scan logging to 1/sec
        self.last_rollover_date = None  # Ensures rollover fires only once per day

        # --- Opening Range Gap Detection (9:15-9:18) ---
        self.opening_range = {'first': None, 'last': None, 'high': None, 'low': None}
        self.opening_range_processed = False  # True after 9:18 gap logic runs
        self.gap_override_active = False       # True when gap absorbed → suppress SPH/SPL reversal
        self.gap_override_levels = {'high': None, 'low': None}  # Opening range used as temp breakout
        self.gap_sl_locked = None              # 'SPH' or 'SPL' when opening-range SL must not loosen
        self.gap_sl_locked_value = None        # Floor (SHORT) or ceiling (LONG) for locked SL

        # --- Watchdog / Heartbeat ---
        self.last_tick_time = datetime.datetime.now()
        self.last_heartbeat_time = datetime.datetime.now()
        self.ticker_connected = False
        self._ticker_connecting = False
        self._ticker_intentional_close = False
        self._eod_disconnect_date = None
        self._morning_connect_date = None
        self._last_connect_attempt = None
        self._morning_login_date = None
        self._morning_login_failed = False
        self._morning_login_last_attempt = None
        self.pending_live_direction = None

        # --- Profit Booking State ---
        self.last_book_time = None  # last time profit-book check fired

        # --- Structural signal timing ---
        self.last_reversal_time = None
        self.last_15min_signal_boundary = None
        self.pivot_meta = {}

        # --- Telegram (from bot_secrets.env) ---
        load_secrets()
        self.telegram_bot_token = require_env("TELEGRAM_BOT_TOKEN")
        self.telegram_chat_id = require_env("TELEGRAM_CHAT_ID")
        self.last_update_id = 0
        self.stop_signal_received = False

        # --- Boot Sequence ---
        if self.mode != 'BACKTEST':
            self.setup_kite()
            self.setup_ticker()
            self.load_state()

    # ═══════════════════════════════════════════════
    # KITE CONNECTION & INSTRUMENT SETUP
    # ═══════════════════════════════════════════════

    def setup_kite(self):
        """Connect to Kite API and validate session (auto-login if token expired)."""
        try:
            self.access_token = ensure_access_token()
            self.kite = KiteConnect(api_key=self.api_key)
            self.kite.set_access_token(self.access_token)
            self.kite.profile()  # Validate session
            logger.info("✅ Kite Connected & Validated")
            self.setup_instruments()
        except KiteLoginError as e:
            logger.error(f"Kite auto-login failed: {e}")
            self.send_telegram(f"❌ Kite auto-login failed: {e}")
        except Exception as e:
            # Fallback: stale file on disk but ensure_access_token already retried login.
            try:
                token = read_access_token()
                if token:
                    self.access_token = token
                    self.kite = KiteConnect(api_key=self.api_key)
                    self.kite.set_access_token(self.access_token)
                    self.kite.profile()
                    logger.info("✅ Kite Connected & Validated")
                    self.setup_instruments()
                    return
            except Exception:
                pass
            logger.error(f"Kite Connection Failed: {e}")
            self.send_telegram(f"❌ Kite Connection Failed: {e}")

    def setup_instruments(self):
        """
        Fetch NFO instruments and auto-select NEXT WEEK expiry.
        
        NIFTY weekly expiry = Tuesday (Mon if Tue holiday, Fri if Mon+Tue holiday).
        We always want "next week" expiry to avoid theta decay.
        
        Logic: If the nearest expiry is ≤ 4 days away, skip it → pick the next one.
        This ensures rollover happens by Friday before Tuesday expiry:
          - Fri (4 days to Tue): ≤4 → skip → picks following week ✓
          - Sat (3 days): ≤4 → skip ✓
          - Sun (2 days): ≤4 → skip ✓
          - Mon (1 day):  ≤4 → skip ✓
          - Tue (expiry):  ≤4 → skip ✓
          - Wed (6 days to next Tue): >4 → use it ✓ (IS next week)
          - Thu (5 days): >4 → use it ✓
        """
        try:
            logger.info("Fetching NFO Instruments...")
            instruments = self.kite.instruments("NFO")
            df = pd.DataFrame(instruments)
            df = df[(df['name'] == 'NIFTY') & (df['segment'] == 'NFO-OPT')]
            df['expiry'] = pd.to_datetime(df['expiry']).dt.date

            today = datetime.date.today()
            valid_expiries = sorted(df[df['expiry'] >= today]['expiry'].unique())

            if not valid_expiries:
                logger.error("No valid expiries found!")
                return

            # --- NEXT-WEEK EXPIRY SELECTION (Tuesday expiry cycle) ---
            # Skip nearest if it's within 4 days (i.e., we're Fri-Tue of that week)
            if len(valid_expiries) >= 2 and (valid_expiries[0] - today).days <= 4:
                self.expiry_date = valid_expiries[1]
            else:
                self.expiry_date = valid_expiries[0]

            logger.info(f"✅ Selected Next-Week Expiry: {self.expiry_date} (Today: {today}, Nearest available: {valid_expiries[0]})")
            self.instruments_df = df[df['expiry'] == self.expiry_date]

        except Exception as e:
            logger.error(f"Instrument Setup Failed: {e}")

    def get_option_symbol(self, strike, type_):
        """Get tradingsymbol and instrument_token for a specific strike & type (CE/PE)."""
        try:
            row = self.instruments_df[
                (self.instruments_df['strike'] == strike) &
                (self.instruments_df['instrument_type'] == type_)
            ]
            if not row.empty:
                return row.iloc[0]['tradingsymbol'], row.iloc[0]['instrument_token']
        except:
            pass
        return None, None

    def get_option_ltp(self, symbol):
        """Get last traded price for an option symbol."""
        try:
            ltp_data = self.kite.ltp([f"NFO:{symbol}"])
            if ltp_data and f"NFO:{symbol}" in ltp_data:
                return ltp_data[f"NFO:{symbol}"]['last_price']
        except Exception as e:
            logger.error(f"LTP Fetch Error ({symbol}): {e}")
        return 0

    # ═══════════════════════════════════════════════
    # WEBSOCKET TICKER
    # ═══════════════════════════════════════════════

    def setup_ticker(self):
        """Initialize KiteTicker websocket."""
        try:
            self.kws = KiteTicker(self.api_key, self.access_token)
            self.kws.on_ticks = self.on_ticks
            self.kws.on_connect = self.on_connect
            self.kws.on_close = self.on_close
            self.kws.on_error = self.on_error
        except Exception as e:
            logger.error(f"Ticker Setup Failed: {e}")

    def on_connect(self, kws, response):
        self.ticker_connected = True
        self._ticker_connecting = False
        if not hasattr(self, "_session_refresh_date"):
            self._session_refresh_date = None
        self._session_refresh_date = datetime.datetime.now().date()
        logger.info("Ticker Connected. Subscribing to NIFTY...")
        kws.subscribe([self.instrument_token])
        kws.set_mode(kws.MODE_FULL, [self.instrument_token])

    def on_error(self, kws, code, reason):
        self._ticker_connecting = False
        self.ticker_connected = False
        logger.error(f"Ticker error: {code} - {reason}")

    def on_close(self, kws, code, reason):
        self.ticker_connected = False
        self._ticker_connecting = False
        if self._ticker_intentional_close:
            logger.info(f"Ticker closed (scheduled): {code} - {reason}")
            return
        logger.error(f"Ticker Closed: {code} - {reason}")

    def is_trading_day(self, dt=None):
        """Mon–Fri (NSE cash session days; no holiday calendar)."""
        dt = dt or datetime.datetime.now()
        return dt.weekday() < 5

    def is_ticker_session_hours(self, dt=None):
        """Weekday 8:00–15:30 IST — websocket should be connected."""
        dt = dt or datetime.datetime.now()
        if not self.is_trading_day(dt):
            return False
        t = dt.time()
        connect_at = datetime.time(*TICKER_CONNECT_TIME)
        disconnect_at = datetime.time(*TICKER_DISCONNECT_TIME)
        return connect_at <= t < disconnect_at

    def connect_ticker(self, reason=""):
        """Connect KiteTicker websocket during trading session."""
        if self.mode == 'BACKTEST':
            return
        if self.ticker_connected or self._ticker_connecting:
            return
        self._ticker_connecting = True
        try:
            if self.kws:
                try:
                    self._ticker_intentional_close = True
                    self.kws.close()
                except Exception:
                    pass
                self.kws = None
                time.sleep(0.5)

            self.access_token = ensure_access_token()
            if self.kite:
                self.kite.set_access_token(self.access_token)
            self.setup_ticker()
            if not self.kws:
                raise RuntimeError("Ticker setup failed")
            self._ticker_intentional_close = False
            logger.info(f"Connecting ticker — {reason}")
            self.kws.connect(threaded=True)
            self.last_tick_time = datetime.datetime.now()
        except Exception as e:
            self._ticker_connecting = False
            self.ticker_connected = False
            logger.error(f"Ticker connect failed ({reason}): {e}")
            self.send_telegram(f"❌ Ticker connect failed: {e}")

    def disconnect_ticker(self, reason=""):
        """Disconnect websocket (EOD or maintenance); process keeps running."""
        if not self.kws:
            self.ticker_connected = False
            self._ticker_connecting = False
            return
        try:
            self._ticker_intentional_close = True
            self.kws.close()
        except Exception as e:
            logger.warning(f"Ticker disconnect ({reason}): {e}")
        self.ticker_connected = False
        self._ticker_connecting = False

    def _handle_eod_disconnect(self, now):
        """Disconnect websocket after market close; hold positions overnight."""
        today = now.date()
        if self._eod_disconnect_date == today:
            return
        self._eod_disconnect_date = today
        self.save_state()
        pos_desc = (
            f"HELD OVERNIGHT ({self.positions[0]['type']} in {self.positions[0]['symbol']})"
            if self.positions
            else "FLAT"
        )
        connect_h, connect_m = TICKER_CONNECT_TIME
        msg = (
            f"🌙 **MARKET CLOSED (15:30)**\n"
            f"Websocket disconnected until next trading day {connect_h:02d}:{connect_m:02d} AM.\n"
            f"Positions: {pos_desc}\n"
            f"Pivots: SPH={self.sph} | SPL={self.spl}"
        )
        logger.info(msg.replace("**", ""))
        self.send_telegram(msg)
        self.disconnect_ticker("market close 15:30")

    def _ensure_morning_kite_login(self, now) -> bool:
        """One coordinated login at/after 8 AM (retries, no parallel logins)."""
        if not self.is_trading_day(now) or now.hour < TICKER_CONNECT_TIME[0]:
            return False

        today = now.date()
        if self._morning_login_date == today:
            token = read_access_token()
            return bool(token and token_is_valid(self.api_key, token))

        if self._morning_login_last_attempt:
            elapsed = (now - self._morning_login_last_attempt).total_seconds()
            if self._morning_login_failed and elapsed < 300:
                return False

        self._morning_login_last_attempt = now
        try:
            self.access_token = ensure_access_token(force=True, max_retries=8)
            if self.kite:
                self.kite.set_access_token(self.access_token)
            self._morning_login_date = today
            self._morning_login_failed = False
            self._session_refresh_date = today
            logger.info("✅ Morning Kite login succeeded")
            self.send_telegram("✅ Morning Kite login succeeded (8 AM IST)")
            return True
        except Exception as e:
            self._morning_login_failed = True
            logger.error(f"Morning Kite login failed: {e}")
            self.send_telegram(
                f"❌ Morning Kite login failed: {e}\nWill retry in ~5 min."
            )
            return False

    def _maybe_connect_morning_session(self, now):
        """Login at 8 AM, then connect websocket (retries until 10 AM)."""
        if not self.is_ticker_session_hours(now) or self.ticker_connected:
            return
        if now.hour < TICKER_CONNECT_TIME[0]:
            return

        if not self._ensure_morning_kite_login(now):
            return

        if self._ticker_connecting:
            # Stuck handshake — allow retry after 45s
            if (
                self._last_connect_attempt
                and (now - self._last_connect_attempt).total_seconds() < 45
            ):
                return
            self._ticker_connecting = False
            if self.kws:
                try:
                    self._ticker_intentional_close = True
                    self.kws.close()
                except Exception:
                    pass
                self.kws = None
        if now.hour < TICKER_CONNECT_TIME[0]:
            return
        if (
            self._last_connect_attempt
            and (now - self._last_connect_attempt).total_seconds() < 120
        ):
            return
        self._last_connect_attempt = now
        if self._morning_connect_date != now.date():
            self._morning_connect_date = now.date()
        self.connect_ticker("morning session")

    def on_ticks(self, kws, ticks):
        for tick in ticks:
            if tick['instrument_token'] == self.instrument_token:
                self.process_tick(tick)

    # ═══════════════════════════════════════════════
    # TICK PROCESSING
    # Pivots refresh on each 15-min boundary. Structural SPH/SPL signals use the
    # last closed 15-min candle when EXECUTE_ON_15MIN_CLOSE is True; otherwise ticks.
    # Opening-range override and profit-book still run on ticks.
    # ═══════════════════════════════════════════════

    def process_tick(self, tick):
        """Process each incoming NIFTY tick."""
        ltp = tick['last_price']
        tick_dt = datetime.datetime.now()
        self.last_tick_time = tick_dt

        # --- OPENING RANGE COLLECTION (9:15 - 9:18) ---
        # Collect ticks during the first 3 minutes to detect gap direction.
        # At 9:18, use the micro-candle to decide if a gap is confirmed or absorbed.
        if tick_dt.time() < datetime.time(9, 15):
            return  # Pre-open auction, ignore

        if tick_dt.time() < datetime.time(9, 18):
            # Collect opening range ticks (don't execute yet)
            if self.opening_range['first'] is None:
                self.opening_range = {
                    'first': ltp, 'last': ltp,
                    'high': ltp, 'low': ltp
                }
                logger.info(f"📊 Opening Range Started: First tick = {ltp}")
            else:
                self.opening_range['last'] = ltp
                self.opening_range['high'] = max(self.opening_range['high'], ltp)
                self.opening_range['low'] = min(self.opening_range['low'], ltp)
            return  # Don't execute until 9:18

        # --- AT 9:18: PROCESS GAP DETECTION (once) ---
        if not self.opening_range_processed:
            self.opening_range_processed = True
            self._process_opening_gap(tick_dt, ltp)

        # ── STEP 1: CANDLE MANAGEMENT (runs first, keeps pivots fresh) ──
        minute_key = tick_dt.replace(second=0, microsecond=0)

        if self.current_minute is None:
            self.current_minute = minute_key
            self.current_candle = {'open': ltp, 'high': ltp, 'low': ltp, 'close': ltp}

        elif minute_key > self.current_minute:
            # Minute closed — store candle
            new_row = pd.DataFrame([{
                'open': self.current_candle['open'],
                'high': self.current_candle['high'],
                'low': self.current_candle['low'],
                'close': self.current_candle['close'],
                'volume': 0
            }], index=[self.current_minute])

            self.df_1min = pd.concat([self.df_1min, new_row])
            self.df_1min.sort_index(inplace=True)

            # Start new candle
            self.current_minute = minute_key
            self.current_candle = {'open': ltp, 'high': ltp, 'low': ltp, 'close': ltp}

        else:
            # Update current candle
            self.current_candle['high'] = max(self.current_candle['high'], ltp)
            self.current_candle['low'] = min(self.current_candle['low'], ltp)
            self.current_candle['close'] = ltp

        # ── 15-MIN PIVOT REFRESH (re-fetch from Kite API at each boundary) ──
        # Instead of stitching candles from ticks (which can drift or miss data),
        # we re-fetch native 15-min candles from Kite at each 15-min boundary.
        min_15_key = tick_dt.replace(second=0, microsecond=0, minute=(tick_dt.minute // 15) * 15)

        if self.current_15min_key is None:
            self.current_15min_key = min_15_key

        elif min_15_key > self.current_15min_key:
            # 15-min boundary crossed — re-fetch candles from Kite API
            self.current_15min_key = min_15_key

            try:
                now = datetime.datetime.now()
                from_date = now - datetime.timedelta(days=10)
                records = self.kite.historical_data(
                    self.instrument_token, from_date, now, "15minute"
                )
                if records:
                    df_hist = pd.DataFrame(records)
                    df_hist['date'] = pd.to_datetime(df_hist['date']).dt.tz_localize(None)
                    df_hist.set_index('date', inplace=True)
                    self.df_15min = df_hist
                    self.recalculate_pivots()
                    if EXECUTE_ON_15MIN_CLOSE:
                        self._run_15min_bar_signals(tick_dt)
            except Exception as e:
                logger.error(f"15-min refresh error: {e}")

        # ── STEP 2: EXPIRY DAY EXIT (one-time trigger at 15:15) ──
        today = datetime.date.today()
        if self.expiry_date and today == self.expiry_date:
            if tick_dt.hour == 15 and tick_dt.minute >= 15:
                expiry_key = tick_dt.strftime('%Y-%m-%d-EXPIRY')
                if self.last_exec_15min != expiry_key:
                    self.handle_expiry_exit(tick_dt, ltp)
                    self.last_exec_15min = expiry_key
                    return

        # ── STEP 2.5: PRE-EXPIRY AUTO-ROLLOVER (at 15:20) ──
        # If held options expire within 4 days, roll to next week.
        # Triggers at 3:20 PM (EOD) so we trade the full day with current expiry.
        # Better pricing and avoids morning volatility affecting the roll.
        if self.positions and self.last_rollover_date != str(today):
            if tick_dt.hour == 15 and tick_dt.minute >= 20:
                pos_expiry_str = self.positions[0].get('expiry', '')
                if pos_expiry_str:
                    try:
                        pos_expiry = datetime.date.fromisoformat(pos_expiry_str)
                        days_to_expiry = (pos_expiry - today).days
                        if 0 < days_to_expiry <= 4:
                            logger.info(f"📅 PRE-EXPIRY ROLLOVER (15:20): Position expiry {pos_expiry} is {days_to_expiry} days away. Rolling over...")
                            self.handle_expiry_rollover(tick_dt, ltp)
                            self.last_rollover_date = str(today)
                            return
                    except ValueError:
                        pass

        # ── STEP 3: END OF DAY — DISCONNECT WEBSOCKET AT 15:30 ──
        # Positions are kept open (positional strategy — held overnight).
        disconnect_h, disconnect_m = TICKER_DISCONNECT_TIME
        if tick_dt.hour == disconnect_h and tick_dt.minute >= disconnect_m:
            if self._eod_disconnect_date != tick_dt.date():
                self._handle_eod_disconnect(tick_dt)
            return

        # ── STEP 3.5: PROFIT BOOK + STRUCTURAL RE-ENTRY ──
        # If the active option premium has gained ≥ PROFIT_BOOK_PCT, exit and
        # only re-enter same direction if price is still beyond the active level.
        if PROFIT_BOOK_ENABLED and self.positions:
            self.check_profit_book(tick_dt, ltp)

        # ── STEP 4: BREAKOUT / OR OVERRIDE ──
        # Structural SPH/SPL on 15-min close only when EXECUTE_ON_15MIN_CLOSE; else every tick.
        allow_structural = not EXECUTE_ON_15MIN_CLOSE
        self.execute_logic(tick_dt, ltp, allow_structural=allow_structural)

    # ═══════════════════════════════════════════════
    # OPENING GAP DETECTION (9:15 - 9:18)
    # ═══════════════════════════════════════════════

    def fetch_opening_range_from_history(self, t):
        """
        If the bot starts after 09:18, live ticks never fill opening_range.
        Rebuild 9:15–9:17 from NIFTY 1-minute candles (same window as the first three minutes).
        """
        if not self.kite:
            return False
        today = t.date() if hasattr(t, "date") else datetime.date.today()
        try:
            from_dt = datetime.datetime.combine(today, datetime.time(9, 0))
            to_dt = max(
                datetime.datetime.combine(today, datetime.time(9, 20)),
                t,
            )
            records = self.kite.historical_data(
                self.instrument_token, from_dt, to_dt, "minute"
            )
        except Exception as e:
            logger.error(f"Opening range history fetch failed: {e}")
            return False
        if not records:
            return False
        df = pd.DataFrame(records)
        df["date"] = pd.to_datetime(df["date"])
        if len(df) > 0 and df["date"].iloc[0].tzinfo is not None:
            df["date"] = df["date"].dt.tz_convert("Asia/Kolkata").dt.tz_localize(None)
        df.set_index("date", inplace=True)
        df.sort_index(inplace=True)
        try:
            seg = df.between_time("09:15", "09:17", inclusive="both")
        except Exception:
            mask = (df.index.time >= datetime.time(9, 15)) & (
                df.index.time <= datetime.time(9, 17)
            )
            seg = df.loc[mask]
        if seg.empty:
            logger.warning(
                "Opening range history: no 9:15–9:17 minute bars (holiday or no data yet)"
            )
            return False
        first_open = float(seg.iloc[0]["open"])
        last_close = float(seg.iloc[-1]["close"])
        hi = float(seg["high"].max())
        lo = float(seg["low"].min())
        self.opening_range = {
            "first": first_open,
            "last": last_close,
            "high": hi,
            "low": lo,
            "hydrated": True,
        }
        logger.info(
            f"📊 Opening range hydrated from Kite 1m (9:15–9:17): "
            f"first={first_open} last={last_close} H={hi} L={lo}"
        )
        return True

    def _process_opening_gap(self, t, current_ltp):
        """
        At 9:18, analyze the opening range (9:15-9:18 ticks) to handle gap opens.

        FIRST CHECK: If gap is IN OUR FAVOR (profitable), skip OR and let it ride.
        - SHORT + gap DOWN = in our favor → continue SHORT, normal pivots resume
        - LONG + gap UP = in our favor → continue LONG, normal pivots resume

        ONLY IF gap is AGAINST us or we're FLAT:
        - If market opened with a gap through SPH or SPL:
            - Check micro-candle direction (first tick vs last tick)
            - Gap DOWN + ticks still falling (last < first) → confirmed, enter SHORT
            - Gap DOWN + ticks recovering (last > first) → absorbed, skip (wait for normal breakout)
            - Gap UP + ticks still rising (last > first) → confirmed, enter LONG
            - Gap UP + ticks falling back (last < first) → absorbed, skip
        - If no gap (opened between SPH and SPL): do nothing, normal logic handles it.
        """
        SPH = self.sph
        SPL = self.spl
        orng = self.opening_range

        if pd.isna(SPH) or pd.isna(SPL):
            logger.info("📊 Opening Range: No pivots — skipping gap detection")
            return

        if orng["first"] is None:
            if not self.fetch_opening_range_from_history(t):
                logger.info(
                    "📊 Opening Range: No live ticks and history unavailable — skipping gap detection"
                )
                return

        orng = self.opening_range
        hydrated = orng.get("hydrated", False)
        # Late boot: use end-of-window price for gap confirm, not LTP hours later.
        decision_ltp = orng["last"] if hydrated else current_ltp

        first = orng['first']
        last = orng['last']
        orng_high = orng['high']
        orng_low = orng['low']

        logger.info(
            f"📊 Opening Range Analysis:\n"
            f"  First tick: {first} | Last tick: {last}\n"
            f"  Range: H={orng_high} L={orng_low}\n"
            f"  Pivots: SPH={SPH} | SPL={SPL}\n"
            f"  Micro-trend: {'BEARISH ↓' if last < first else 'BULLISH ↑' if last > first else 'FLAT →'}"
        )

        # ── GAP IN OUR FAVOR: Let it ride! ──
        # If the gap makes our existing position MORE profitable, skip opening range
        # and let normal SPH/SPL pivots resume naturally.
        if self.positions:
            pos_type = self.positions[0]['type']

            # SHORT + gap DOWN = already in profit → tighten SL to opening high
            if pos_type == 'SHORT' and orng_low < SPL:
                self._apply_opening_sl('SHORT', orng_high, orng_low, lock=True)
                logger.info(
                    f"🎯 GAP IN OUR FAVOR: SHORT + Gap Down. "
                    f"Opening low ({orng_low}) < SPL ({SPL}). "
                    f"Continue SHORT — SL set to opening high ({orng_high})."
                )
                self.send_telegram(
                    f"🎯 **GAP IN OUR FAVOR**\n"
                    f"Position: SHORT ✅\n"
                    f"Gap: DOWN (Open {orng_low} < SPL {SPL})\n"
                    f"Action: Continue SHORT\n"
                    f"🛑 New SL (SPH): {self.sph}"
                )
                return

            # LONG + gap UP = already in profit → tighten SL to opening low
            if pos_type == 'LONG' and orng_high > SPH:
                self._apply_opening_sl('LONG', orng_high, orng_low, lock=True)
                logger.info(
                    f"🎯 GAP IN OUR FAVOR: LONG + Gap Up. "
                    f"Opening high ({orng_high}) > SPH ({SPH}). "
                    f"Continue LONG — SL set to opening low ({orng_low})."
                )
                self.send_telegram(
                    f"🎯 **GAP IN OUR FAVOR**\n"
                    f"Position: LONG ✅\n"
                    f"Gap: UP (Open {orng_high} > SPH {SPH})\n"
                    f"Action: Continue LONG\n"
                    f"🛑 New SL (SPL): {self.spl}"
                )
                return

        # ── GAP AGAINST US or FLAT: Use Opening Range logic ──

        # --- GAP DOWN: Opening range went below SPL ---
        if orng_low < SPL:
            if decision_ltp < orng_low:
                # 9:18 tick breaks below the 9:15-9:17 LOW → sellers pushing new lows → SHORT
                logger.info(f"⚡ GAP DOWN CONFIRMED: decision LTP ({decision_ltp}) broke below opening low ({orng_low}). Entering SHORT.")
                self.send_telegram(
                    f"⚡ **GAP DOWN CONFIRMED**\n"
                    f"Opening Range: H={orng_high} L={orng_low}\n"
                    f"Decision LTP: {decision_ltp} < Opening Low ({orng_low})\n"
                    f"SPL: {SPL} broken + new low\n"
                    f"Action: Reverse to SHORT"
                )
                # Reverse carried position if needed
                if self.positions and self.positions[0]['type'] == 'LONG':
                    self._exit_position(t, decision_ltp, self.positions[0], 'GAP_DOWN_CONFIRMED')
                if not self.positions:
                    self._enter_position(t, decision_ltp, 'SHORT', 'GAP_DOWN_CONFIRMED')
            else:
                # 9:18 tick holds above the opening low → gap absorbed
                # KEEP current position, use opening range as temp breakout levels
                self.gap_override_active = True
                self.gap_override_levels = {'high': orng_high, 'low': orng_low}
                if self.positions:
                    self._apply_opening_sl(
                        self.positions[0]['type'], orng_high, orng_low, lock=True
                    )
                logger.info(f"🛡️ GAP DOWN ABSORBED: Keeping current position. Opening range override active (H={orng_high}, L={orng_low}).")
                kept = (
                    'SHORT' if self.positions and self.positions[0]['type'] == 'SHORT'
                    else 'LONG' if self.positions else 'FLAT'
                )
                sl_line = (
                    f"🛑 New SL (SPL): {self.spl}" if kept == 'LONG'
                    else f"🛑 New SL (SPH): {self.sph}" if kept == 'SHORT'
                    else ""
                )
                self.send_telegram(
                    f"🛡️ **GAP DOWN ABSORBED**\n"
                    f"Opening Range: H={orng_high} L={orng_low}\n"
                    f"Decision LTP: {decision_ltp} > Opening Low ({orng_low})\n"
                    f"Keeping {kept}\n"
                    f"🔀 Temp breakout levels: H={orng_high} L={orng_low}\n"
                    f"{sl_line}"
                )

        # --- GAP UP: Opening range went above SPH ---
        elif orng_high > SPH:
            if decision_ltp > orng_high:
                # 9:18 tick breaks above the 9:15-9:17 HIGH → buyers pushing new highs → LONG
                logger.info(f"⚡ GAP UP CONFIRMED: decision LTP ({decision_ltp}) broke above opening high ({orng_high}). Entering LONG.")
                self.send_telegram(
                    f"⚡ **GAP UP CONFIRMED**\n"
                    f"Opening Range: H={orng_high} L={orng_low}\n"
                    f"Decision LTP: {decision_ltp} > Opening High ({orng_high})\n"
                    f"SPH: {SPH} broken + new high\n"
                    f"Action: Reverse to LONG"
                )
                # Reverse carried position if needed
                if self.positions and self.positions[0]['type'] == 'SHORT':
                    self._exit_position(t, decision_ltp, self.positions[0], 'GAP_UP_CONFIRMED')
                if not self.positions:
                    self._enter_position(t, decision_ltp, 'LONG', 'GAP_UP_CONFIRMED')
            else:
                # 9:18 tick below the opening high → gap absorbed
                # KEEP current position, use opening range as temp breakout levels
                self.gap_override_active = True
                self.gap_override_levels = {'high': orng_high, 'low': orng_low}
                if self.positions:
                    self._apply_opening_sl(
                        self.positions[0]['type'], orng_high, orng_low, lock=True
                    )
                logger.info(f"🛡️ GAP UP ABSORBED: Keeping current position. Opening range override active (H={orng_high}, L={orng_low}).")
                kept = (
                    'SHORT' if self.positions and self.positions[0]['type'] == 'SHORT'
                    else 'LONG' if self.positions else 'FLAT'
                )
                sl_line = (
                    f"🛑 New SL (SPH): {self.sph}" if kept == 'SHORT'
                    else f"🛑 New SL (SPL): {self.spl}" if kept == 'LONG'
                    else ""
                )
                self.send_telegram(
                    f"🛡️ **GAP UP ABSORBED**\n"
                    f"Opening Range: H={orng_high} L={orng_low}\n"
                    f"Decision LTP: {decision_ltp} < Opening High ({orng_high})\n"
                    f"Keeping {kept}\n"
                    f"🔀 Temp breakout levels: H={orng_high} L={orng_low}\n"
                    f"{sl_line}"
                )

        # --- NO GAP: Opened between SPH and SPL ---
        else:
            logger.info(f"📊 No gap detected. Opened within SPH-SPL range. Normal breakout logic applies.")
            self.send_telegram(
                f"📊 **NO GAP (9:18)**\n"
                f"Opening Range: H={orng_high} L={orng_low}\n"
                f"SPH={SPH} | SPL={SPL}\n"
                f"Opened inside range — waiting for SPH/SPL breakout."
            )

    def _clear_gap_sl_lock(self):
        """Clear opening-range SL lock (on exit or reversal)."""
        self.gap_sl_locked = None
        self.gap_sl_locked_value = None

    def _apply_opening_sl(self, pos_type, or_high, or_low, lock=True, notify=True):
        """
        Set trailing SL from opening range.

        SHORT → SPH at opening high; LONG → SPL at opening low.
        When lock=True, pivot recalc cannot loosen below this level.
        """
        if pos_type == 'SHORT':
            if pd.isna(self.spl) or or_high <= self.spl:
                logger.warning(
                    f"⚠️ Skipping OR SPH update: OR high ({or_high}) <= SPL ({self.spl})"
                )
                return False
            old_sph = self.sph
            self.sph = or_high
            if lock:
                self.gap_sl_locked = 'SPH'
                self.gap_sl_locked_value = float(or_high)
            if notify and (pd.isna(old_sph) or old_sph != self.sph):
                self.send_sl_update(self.sph, 'SHORT')
            return True

        if pos_type == 'LONG':
            if pd.isna(self.sph) or or_low >= self.sph:
                logger.warning(
                    f"⚠️ Skipping OR SPL update: OR low ({or_low}) >= SPH ({self.sph})"
                )
                return False
            old_spl = self.spl
            self.spl = or_low
            if lock:
                self.gap_sl_locked = 'SPL'
                self.gap_sl_locked_value = float(or_low)
            if notify and (pd.isna(old_spl) or old_spl != self.spl):
                self.send_sl_update(self.spl, 'LONG')
            return True

        return False

    def _maybe_release_gap_sl_lock(self, new_sph, new_spl):
        """
        Release opening-range SL lock when a new intraday pivot supersedes it.

        SHORT: new SPH from today, below opening high → trail normal 15m SPH.
        LONG:  new SPL from today, above opening low  → trail normal 15m SPL.
        """
        if not self.gap_sl_locked or self.gap_sl_locked_value is None:
            return False
        if new_sph is None or new_spl is None or new_sph <= new_spl:
            return False

        meta = self.pivot_meta or {}
        today = datetime.date.today()
        locked = float(self.gap_sl_locked_value)

        if self.gap_sl_locked == 'SPH':
            sph_time = meta.get('raw_sph_time')
            if sph_time is None or sph_time.date() != today:
                return False
            if new_sph >= locked:
                return False
            logger.info(
                f"🔓 Gap SPH lock released: intraday SPH {new_sph} "
                f"< opening high {locked} (bar {sph_time.strftime('%H:%M')})"
            )
            self.send_telegram(
                f"🔓 **GAP SL LOCK RELEASED**\n"
                f"New intraday SPH: {new_sph}\n"
                f"(was locked at opening high {locked})\n"
                f"Normal pivot trailing resumes."
            )
            self._clear_gap_sl_lock()
            self.save_state()
            return True

        if self.gap_sl_locked == 'SPL':
            spl_time = meta.get('raw_spl_time')
            if spl_time is None or spl_time.date() != today:
                return False
            if new_spl <= locked:
                return False
            logger.info(
                f"🔓 Gap SPL lock released: intraday SPL {new_spl} "
                f"> opening low {locked} (bar {spl_time.strftime('%H:%M')})"
            )
            self.send_telegram(
                f"🔓 **GAP SL LOCK RELEASED**\n"
                f"New intraday SPL: {new_spl}\n"
                f"(was locked at opening low {locked})\n"
                f"Normal pivot trailing resumes."
            )
            self._clear_gap_sl_lock()
            self.save_state()
            return True

        return False

    def _enforce_gap_sl_lock(self):
        """Prevent pivot recalc from loosening a gap-based SL (while lock is active)."""
        if self.gap_sl_locked == 'SPH' and self.gap_sl_locked_value is not None and not pd.isna(self.sph):
            self.sph = max(self.sph, self.gap_sl_locked_value)
        elif self.gap_sl_locked == 'SPL' and self.gap_sl_locked_value is not None and not pd.isna(self.spl):
            self.spl = min(self.spl, self.gap_sl_locked_value)

    def _apply_session_sl_cap(self):
        """Optionally cap trailing SPH/SPL to today's session high/low (legacy behavior)."""
        if not SESSION_SL_CAP_ENABLED or not self.positions or self.df_15min.empty:
            return
        pos_type = self.positions[0]['type']
        today = datetime.date.today()
        today_data = self.df_15min[self.df_15min.index.date == today]
        if today_data.empty:
            return
        today_high = today_data['high'].max()
        today_low = today_data['low'].min()

        if pos_type == 'SHORT' and not pd.isna(self.sph) and self.sph > today_high:
            if not pd.isna(self.spl) and today_high > self.spl:
                self.sph = today_high
            else:
                logger.warning(
                    f"⚠️ Skipping SPH cap: today_high ({today_high}) <= SPL ({self.spl})"
                )

        if pos_type == 'LONG' and not pd.isna(self.spl) and self.spl < today_low:
            if not pd.isna(self.sph) and today_low < self.sph:
                self.spl = today_low
            else:
                logger.warning(
                    f"⚠️ Skipping SPL cap: today_low ({today_low}) >= SPH ({self.sph})"
                )

    def _select_active_pivots(self, df_raw, last_bar):
        """
        Pick latest confirmed SPH and SPL independently (avoids stale paired ffill).
        Falls back to last historically valid pair if levels invert.
        """
        recent = df_raw.loc[:last_bar]
        sph_pts = recent['sph'].dropna()
        spl_pts = recent['spl'].dropna()
        meta = {
            'last_bar': last_bar,
            'raw_sph': float(sph_pts.iloc[-1]) if not sph_pts.empty else None,
            'raw_spl': float(spl_pts.iloc[-1]) if not spl_pts.empty else None,
            'raw_sph_time': sph_pts.index[-1] if not sph_pts.empty else None,
            'raw_spl_time': spl_pts.index[-1] if not spl_pts.empty else None,
            'source': 'independent',
        }

        if not sph_pts.empty and not spl_pts.empty:
            cand_sph = float(sph_pts.iloc[-1])
            cand_spl = float(spl_pts.iloc[-1])
            if cand_sph > cand_spl:
                return cand_sph, cand_spl, meta

        df_ff = df_raw.copy()
        df_ff['sph'] = df_ff['sph'].ffill()
        df_ff['spl'] = df_ff['spl'].ffill()
        recent_ff = df_ff.loc[:last_bar]
        valid_pairs = recent_ff[
            recent_ff['sph'].notna()
            & recent_ff['spl'].notna()
            & (recent_ff['sph'] > recent_ff['spl'])
        ]
        if not valid_pairs.empty:
            row = valid_pairs.iloc[-1]
            meta['source'] = 'ffill_pair_fallback'
            meta['fallback_time'] = valid_pairs.index[-1]
            return float(row['sph']), float(row['spl']), meta

        return None, None, meta

    def _notify_pivot_refresh(self, old_sph, old_spl):
        if not PIVOT_TELEGRAM_ALERTS:
            return
        meta = self.pivot_meta or {}
        raw_sph = meta.get('raw_sph')
        raw_spl = meta.get('raw_spl')
        last_bar = meta.get('last_bar')
        bar_str = last_bar.strftime('%H:%M') if last_bar is not None else '—'
        changed = (
            (not pd.isna(self.sph) and self.sph != old_sph)
            or (not pd.isna(self.spl) and self.spl != old_spl)
        )
        if not changed and old_sph is not None:
            return
        self.send_telegram(
            f"📐 **PIVOTS UPDATED** ({bar_str})\n"
            f"Active: SPH={self.sph} | SPL={self.spl}\n"
            f"Latest raw: SPH={raw_sph} | SPL={raw_spl}\n"
            f"Source: {meta.get('source', '—')}"
        )

    def _run_15min_bar_signals(self, tick_dt):
        """Run structural SPH/SPL logic once per 15-min boundary on last candle close."""
        if self.df_15min.empty:
            return
        boundary_key = tick_dt.replace(
            second=0, microsecond=0, minute=(tick_dt.minute // 15) * 15
        )
        if self.last_15min_signal_boundary == boundary_key:
            return
        self.last_15min_signal_boundary = boundary_key

        last_row = self.df_15min.iloc[-1]
        bar_close = float(last_row['close'])
        bar_time = self.df_15min.index[-1]
        logger.info(
            f"📊 15-Min bar signal @ {bar_time} | Close={bar_close} | "
            f"SPH={self.sph} | SPL={self.spl}"
        )
        self.execute_logic(tick_dt, bar_close, allow_structural=True)

    # ═══════════════════════════════════════════════
    # PIVOT CALCULATION (15-MIN, SPH/SPL ONLY)
    # ═══════════════════════════════════════════════

    def calculate_pivots(self, df):
        """
        Calculate raw SPH and SPL pivot points.
        
        Uses asymmetric pivot detection on continuous data.
        Window: 1 left + 1 center + 3 right bars.
        
        - SPH: center HIGH must be the highest in the window
        - SPL: center LOW must be the lowest in the window
        - Left side uses whatever bars are available (up to 1)
        - Right side requires exactly 3 bars for confirmation
        
        This confirms pivots with future context while reacting faster
        than symmetric 3/3 pivots.
        """
        left_w = 1
        right_w = 3
        highs = df['high'].values
        lows = df['low'].values
        n = len(df)

        sph_vals = [np.nan] * n
        spl_vals = [np.nan] * n

        for i in range(n):
            left_start = max(0, i - left_w)
            right_end = i + right_w + 1
            if right_end > n:
                continue  # Not enough right bars yet

            # SPH: center high must be the highest in the window
            window_highs = highs[left_start:right_end]
            if highs[i] == np.max(window_highs):
                sph_vals[i] = highs[i]

            # SPL: center low must be the lowest in the window
            window_lows = lows[left_start:right_end]
            if lows[i] == np.min(window_lows):
                spl_vals[i] = lows[i]

        df['sph'] = sph_vals
        df['spl'] = spl_vals

        # NO ffill here — done in recalculate_pivots
        return df

    def recalculate_pivots(self):
        """
        Calculate SPH/SPL on the latest closed 15-min candle.

        Uses the most recent confirmed SPH and SPL independently (not a stale
        paired ffill row). Optional session high/low cap when SESSION_SL_CAP_ENABLED.
        """
        try:
            if self.df_15min.empty:
                return

            old_sph = self.sph
            old_spl = self.spl

            self.df_15min.sort_index(inplace=True)
            self.df_15min = self.df_15min[~self.df_15min.index.duplicated(keep='last')]

            df_raw = self.calculate_pivots(self.df_15min.copy())
            df_raw.drop(columns=['trading_date'], inplace=True, errors='ignore')
            last_bar = df_raw.index[-1]

            new_sph, new_spl, meta = self._select_active_pivots(df_raw, last_bar)
            self.pivot_meta = meta

            if new_sph is not None and new_spl is not None and new_sph > new_spl:
                if meta.get('source') == 'ffill_pair_fallback':
                    logger.warning(
                        f"⚠️ Independent pivots inverted at {last_bar}; "
                        f"using pair fallback SPH={new_sph} SPL={new_spl}."
                    )
                self.sph = new_sph
                self.spl = new_spl
            else:
                logger.warning(
                    f"⚠️ No valid pivot pair through {last_bar} "
                    f"(cand SPH={new_sph} SPL={new_spl}). Keeping SPH={old_sph} SPL={old_spl}."
                )

            self._apply_session_sl_cap()
            if not self._maybe_release_gap_sl_lock(new_sph, new_spl):
                self._enforce_gap_sl_lock()
            logger.info(f"📐 Pivots (15-Min) → SPH: {self.sph} | SPL: {self.spl}")

            self._notify_pivot_refresh(old_sph, old_spl)

            if self.positions:
                pos_type = self.positions[0]['type']
                if pos_type == 'LONG':
                    if not pd.isna(self.spl) and not pd.isna(old_spl) and self.spl != old_spl:
                        self.send_sl_update(self.spl, 'LONG')
                elif pos_type == 'SHORT':
                    if not pd.isna(self.sph) and not pd.isna(old_sph) and self.sph != old_sph:
                        self.send_sl_update(self.sph, 'SHORT')

        except Exception as e:
            logger.error(f"Pivot Calculation Error: {e}")

    # ═══════════════════════════════════════════════
    # TRADE EXECUTION LOGIC
    # ═══════════════════════════════════════════════

    def execute_logic(self, t, close, allow_structural=True):
        """
        Main strategy logic — SPH/SPL breakout with always-in reversal.

        allow_structural=False: only opening-range override (tick path); used when
        structural signals run on 15-min candle close instead of every tick.

        SAFETY: Skips all breakouts if SPH <= SPL (inverted pivots).

        If gap_override_active:
            Uses opening range HIGH/LOW as temporary breakout levels.
            Once broken, override clears and normal SPH/SPL resumes.

        Normal:
        If FLAT:   Close > SPH → BUY CE (go Long)
                   Close < SPL → BUY PE (go Short)
        If LONG:   Close < SPL → EXIT CE, BUY PE (reverse to Short)
        If SHORT:  Close > SPH → EXIT PE, BUY CE (reverse to Long)
        """
        SPH = self.sph
        SPL = self.spl

        if pd.isna(SPH) or pd.isna(SPL):
            logger.info(f"⏳ Waiting for pivots... SPH={SPH}, SPL={SPL}")
            return
        if SPH <= SPL:
            logger.warning(
                f"🚫 Invalid pivot state (SPH <= SPL). Skipping trade logic. SPH={SPH}, SPL={SPL}"
            )
            return

        # ── OPENING RANGE OVERRIDE (active after absorbed gap) ──
        # Use opening range HIGH/LOW instead of SPH/SPL until resolved.
        if self.gap_override_active:
            or_high = self.gap_override_levels['high']
            or_low = self.gap_override_levels['low']

            active_pos = self.positions[0] if self.positions else None

            if active_pos:
                entry_type = active_pos['type']

                if entry_type == 'SHORT' and close > or_high:
                    # Price broke above opening range → gap was real after all → reverse
                    logger.info(f"⚡ OPENING RANGE BREAK UP: Close ({close}) > OR High ({or_high}). Override cleared → reversing to LONG.")
                    self.gap_override_active = False
                    self._clear_gap_sl_lock()
                    self.send_telegram(
                        f"⚡ **OPENING RANGE BREAK UP**\n"
                        f"LTP: {close} > Opening High ({or_high})\n"
                        f"Override cleared → Reversing SHORT → LONG"
                    )
                    self._exit_position(t, close, active_pos, 'OR_BREAK_UP')
                    self._apply_opening_sl('LONG', or_high, or_low, lock=False)
                    self._enter_position(t, close, 'LONG', 'OR_BREAK_UP')

                elif entry_type == 'LONG' and close < or_low:
                    # Price broke below opening range → gap was real → reverse
                    logger.info(f"⚡ OPENING RANGE BREAK DOWN: Close ({close}) < OR Low ({or_low}). Override cleared → reversing to SHORT.")
                    self.gap_override_active = False
                    self._clear_gap_sl_lock()
                    self.send_telegram(
                        f"⚡ **OPENING RANGE BREAK DOWN**\n"
                        f"LTP: {close} < Opening Low ({or_low})\n"
                        f"Override cleared → Reversing LONG → SHORT"
                    )
                    self._exit_position(t, close, active_pos, 'OR_BREAK_DOWN')
                    self._apply_opening_sl('SHORT', or_high, or_low, lock=False)
                    self._enter_position(t, close, 'SHORT', 'OR_BREAK_DOWN')

                elif entry_type == 'SHORT' and close < or_low:
                    # SHORT validated — price broke below opening range
                    logger.info(f"✅ SHORT VALIDATED: Close ({close}) < OR Low ({or_low}). Override cleared. Normal SPH/SPL resumes.")
                    self.gap_override_active = False
                    self._apply_opening_sl('SHORT', or_high, or_low, lock=True)
                    self.send_telegram(
                        f"✅ **POSITION VALIDATED**\n"
                        f"LTP: {close} < Opening Low ({or_low})\n"
                        f"SHORT confirmed. SL (SPH): {self.sph}"
                    )

                elif entry_type == 'LONG' and close > or_high:
                    # LONG validated — price broke above opening range
                    logger.info(f"✅ LONG VALIDATED: Close ({close}) > OR High ({or_high}). Override cleared. Normal SPH/SPL resumes.")
                    self.gap_override_active = False
                    self._apply_opening_sl('LONG', or_high, or_low, lock=True)
                    self.send_telegram(
                        f"✅ **POSITION VALIDATED**\n"
                        f"LTP: {close} > Opening High ({or_high})\n"
                        f"LONG confirmed. SL (SPL): {self.spl}"
                    )

            else:
                # Flat during override — use opening range for new entries
                if close > or_high:
                    logger.info(f"⚡ OR BREAK UP (flat): Close ({close}) > OR High ({or_high}). Entering LONG.")
                    self.gap_override_active = False
                    self._enter_position(t, close, 'LONG', 'OR_BREAK_UP')
                elif close < or_low:
                    logger.info(f"⚡ OR BREAK DOWN (flat): Close ({close}) < OR Low ({or_low}). Entering SHORT.")
                    self.gap_override_active = False
                    self._enter_position(t, close, 'SHORT', 'OR_BREAK_DOWN')

            return  # Don't run normal SPH/SPL logic while override is active

        if not allow_structural:
            return

        # ── NORMAL SPH/SPL LOGIC ─────────────────────────────
        active_pos = self.positions[0] if self.positions else None

        # ── EXIT + REVERSAL (Always-In) ──────────────────────
        if active_pos:
            entry_type = active_pos['type']
            do_exit = False
            new_direction = None
            reason = ""

            if entry_type == 'LONG' and close < SPL:
                do_exit = True
                new_direction = 'SHORT'
                reason = 'SPL_BREAK'
                logger.info(f"🔄 REVERSAL: LONG → SHORT | Close ({close}) < SPL ({SPL})")

            elif entry_type == 'SHORT' and close > SPH:
                do_exit = True
                new_direction = 'LONG'
                reason = 'SPH_BREAK'
                logger.info(f"🔄 REVERSAL: SHORT → LONG | Close ({close}) > SPH ({SPH})")

            if do_exit and REVERSAL_COOLDOWN_SEC > 0 and self.last_reversal_time:
                elapsed = (t - self.last_reversal_time).total_seconds()
                if elapsed < REVERSAL_COOLDOWN_SEC:
                    logger.info(
                        f"⏳ Reversal cooldown ({REVERSAL_COOLDOWN_SEC}s, "
                        f"{int(elapsed)}s elapsed) — skipping flip"
                    )
                    do_exit = False

            if do_exit:
                self.last_reversal_time = t
                self._exit_position(t, close, active_pos, reason)
                self._enter_position(t, close, new_direction, reason)

        # ── INITIAL ENTRY (Flat, no position) ────────────────
        else:
            # Throttle scan log to once per second (avoid tick-spam)
            sec_key = t.strftime('%Y-%m-%d %H:%M:%S')
            if self.last_scan_log != sec_key:
                logger.info(f"🔎 Scanning: LTP={close} | SPH={SPH} | SPL={SPL}")
                self.last_scan_log = sec_key

            if close > SPH:
                logger.info(f"⚡ BREAKOUT UP: Close ({close}) > SPH ({SPH})")
                self._enter_position(t, close, 'LONG', 'SPH_BREAK')

            elif close < SPL:
                logger.info(f"⚡ BREAKOUT DOWN: Close ({close}) < SPL ({SPL})")
                self._enter_position(t, close, 'SHORT', 'SPL_BREAK')

    def _enter_position(self, t, spot_price, direction, reason):
        """
        Enter a new position based on ORDER_TYPE config.
        
        Supports: BUY, SELL, SYNTHETIC, ZEBRA
        Each builds a list of order legs, places them, and stores position.
        """
        atm = round(spot_price / 50) * 50
        sl_level = self.spl if direction == 'LONG' else self.sph
        legs = []  # List of {'symbol', 'token', 'strike', 'opt_type', 'side', 'qty', 'price'}

        if self.order_type == 'BUY':
            # BUY 1 ITM option
            if direction == 'LONG':
                strike = atm - (50 * self.itm_offset)
                opt_type = 'CE'
            else:
                strike = atm + (50 * self.itm_offset)
                opt_type = 'PE'
            legs.append({'strike': strike, 'opt_type': opt_type, 'side': 'BUY', 'qty': self.qty})

        elif self.order_type == 'SELL':
            # SELL 1 OTM option
            if direction == 'LONG':
                strike = atm - (50 * self.otm_offset)
                opt_type = 'PE'
            else:
                strike = atm + (50 * self.otm_offset)
                opt_type = 'CE'
            legs.append({'strike': strike, 'opt_type': opt_type, 'side': 'SELL', 'qty': self.qty})

        elif self.order_type == 'SYNTHETIC':
            # Buy ITM + Sell OTM
            if direction == 'LONG':
                legs.append({'strike': atm - (50 * self.itm_offset), 'opt_type': 'CE', 'side': 'BUY', 'qty': self.qty})
                legs.append({'strike': atm - (50 * self.otm_offset), 'opt_type': 'PE', 'side': 'SELL', 'qty': self.qty})
            else:
                legs.append({'strike': atm + (50 * self.itm_offset), 'opt_type': 'PE', 'side': 'BUY', 'qty': self.qty})
                legs.append({'strike': atm + (50 * self.otm_offset), 'opt_type': 'CE', 'side': 'SELL', 'qty': self.qty})

        elif self.order_type == 'ZEBRA':
            # Buy 2x Deep ITM + Sell 1x ATM (Zero Extrinsic Back Ratio)
            zebra_buy_qty = self.qty * 2
            zebra_sell_qty = self.qty
            if direction == 'LONG':
                legs.append({'strike': atm - (50 * self.zebra_deep_itm), 'opt_type': 'CE', 'side': 'BUY', 'qty': zebra_buy_qty})
                legs.append({'strike': atm, 'opt_type': 'CE', 'side': 'SELL', 'qty': zebra_sell_qty})
            else:
                legs.append({'strike': atm + (50 * self.zebra_deep_itm), 'opt_type': 'PE', 'side': 'BUY', 'qty': zebra_buy_qty})
                legs.append({'strike': atm, 'opt_type': 'PE', 'side': 'SELL', 'qty': zebra_sell_qty})

        # Resolve symbols, get prices, and place orders
        filled_legs = []
        for leg in legs:
            symbol, token = self.get_option_symbol(leg['strike'], leg['opt_type'])
            if not symbol:
                logger.error(f"Option Not Found: {leg['strike']} {leg['opt_type']}")
                self.send_telegram(f"❌ Option Not Found: {leg['strike']} {leg['opt_type']}")
                return

            price = self.get_option_ltp(symbol)

            if self.mode == 'LIVE':
                try:
                    self.place_order(symbol, leg['side'], leg['qty'])
                    logger.info(f"✅ LIVE Order: {leg['side']} {symbol} x {leg['qty']}")
                except Exception as e:
                    logger.error(f"❌ Entry Order Failed: {e}")
                    self.send_telegram(f"❌ Entry Order Failed ({leg['side']} {symbol}): {e}")
                    return

            filled_legs.append({
                'symbol': symbol,
                'strike': leg['strike'],
                'opt_type': leg['opt_type'],
                'side': leg['side'],
                'qty': leg['qty'],
                'entry_price': price
            })

        # Store position (primary leg = first leg for PnL tracking)
        primary = filled_legs[0]
        self.positions = [{
            'type': direction,
            'spot_price': spot_price,
            'time': str(t),
            'symbol': primary['symbol'],
            'strike': primary['strike'],
            'opt_type': primary['opt_type'],
            'entry_price': primary['entry_price'],
            'expiry': str(self.expiry_date),
            'order_type': self.order_type,
            'legs': filled_legs
        }]
        self.save_state()

        # ── TELEGRAM: NEW ENTRY ──
        legs_desc = "\n".join([
            f"  {'🟢' if l['side']=='BUY' else '🔴'} {l['side']} {l['symbol']} @ ₹{l['entry_price']:.2f} x {l['qty']}"
            for l in filled_legs
        ])
        entry_msg = (
            f"🚀 **NEW ENTRY: {direction}** ({self.order_type})\n"
            f"Spot: {spot_price}\n"
            f"{legs_desc}\n"
            f"Lots: {self.num_lots} | Expiry: {self.expiry_date}\n"
            f"🛑 Trailing SL: {sl_level}\n"
            f"Reason: {reason}\n"
            f"Time: {t.strftime('%d-%b %H:%M:%S')}"
        )
        logger.info(entry_msg)
        self.send_telegram(entry_msg)

        try:
            os.system('afplay /System/Library/Sounds/Ping.aiff&')
        except:
            pass

    # ═══════════════════════════════════════════════
    # PROFIT BOOKING + STRUCTURAL RE-ENTRY
    # ═══════════════════════════════════════════════

    def _premium_gain_pct(self, pos):
        """
        Return current premium gain as a fraction of entry premium across all legs.
        BUY legs:   pnl = (ltp - entry) * qty   (debit; gain when ltp rises)
        SELL legs:  pnl = (entry - ltp) * qty   (credit; gain when ltp falls)
        Basis = sum(|entry * qty|) — gross premium exposure.
        Returns None if basis is 0/missing or LTPs cannot be fetched.
        """
        legs = pos.get('legs') or []
        if not legs:
            entry = float(pos.get('entry_price', 0) or 0)
            qty = int(pos.get('qty', self.qty) or self.qty)
            symbol = pos.get('symbol')
            if not symbol or entry <= 0 or qty <= 0:
                return None
            ltp = self.get_option_ltp(symbol) or 0
            if ltp <= 0:
                return None
            return ((ltp - entry) * qty) / (entry * qty)

        total_basis = 0.0
        total_pnl = 0.0
        for leg in legs:
            sym = leg.get('symbol')
            entry = float(leg.get('entry_price', 0) or 0)
            qty = int(leg.get('qty', 0) or 0)
            if not sym or entry <= 0 or qty <= 0:
                continue
            ltp = self.get_option_ltp(sym) or 0
            if ltp <= 0:
                continue
            total_basis += entry * qty
            if leg.get('side') == 'BUY':
                total_pnl += (ltp - entry) * qty
            else:
                total_pnl += (entry - ltp) * qty

        if total_basis <= 0:
            return None
        return total_pnl / total_basis

    def check_profit_book(self, t, close):
        """
        If active position's premium gain ≥ PROFIT_BOOK_PCT:
          1. Exit current position (reason: BOOK_60PCT).
          2. If price is still beyond the active level (or the level filter is disabled)
             AND we are before BOOK_REENTRY_CUTOFF, re-enter same direction at fresh 1 ITM.
          3. Otherwise stay flat — normal SPH/SPL breakout logic resumes on later ticks.
        """
        if not self.positions:
            return

        # Cooldown to avoid hammering LTP API on every tick after a check
        if self.last_book_time is not None:
            elapsed = (t - self.last_book_time).total_seconds()
            if elapsed < BOOK_REENTRY_COOLDOWN_SEC:
                return
        self.last_book_time = t

        pos = self.positions[0]
        gain = self._premium_gain_pct(pos)
        if gain is None or gain < PROFIT_BOOK_PCT:
            return

        direction = pos.get('type')
        logger.info(
            f"💰 PROFIT BOOK TRIGGER: gain={gain*100:.1f}% ≥ {PROFIT_BOOK_PCT*100:.0f}% on {direction}. Booking now."
        )
        self.send_telegram(
            f"💰 **BOOK PROFIT @ {gain*100:.0f}%**\n"
            f"Direction: {direction}\n"
            f"Threshold: {PROFIT_BOOK_PCT*100:.0f}% of premium\n"
            f"Action: Exit + check structural re-entry"
        )

        self._exit_position(t, close, pos, f'BOOK_{int(PROFIT_BOOK_PCT*100)}PCT')

        # ── Re-entry decision ──
        cutoff_h, cutoff_m = BOOK_REENTRY_CUTOFF
        if (t.hour, t.minute) >= (cutoff_h, cutoff_m):
            msg = f"⏰ Past re-entry cutoff ({cutoff_h:02d}:{cutoff_m:02d}). Staying flat."
            logger.info(msg)
            self.send_telegram(msg)
            return

        if not REENTRY_REQUIRES_LEVEL_BREAK:
            self._enter_position(t, close, direction, f'BOOK_{int(PROFIT_BOOK_PCT*100)}PCT_REENTRY')
            return

        SPH = self.sph
        SPL = self.spl
        if pd.isna(SPH) or pd.isna(SPL):
            logger.info("Re-entry skipped: pivots NaN.")
            self.send_telegram("⏸ Re-entry skipped — pivots NaN.")
            return

        if direction == 'LONG' and close > SPH:
            self._enter_position(t, close, 'LONG', f'BOOK_{int(PROFIT_BOOK_PCT*100)}PCT_REENTRY')
        elif direction == 'SHORT' and close < SPL:
            self._enter_position(t, close, 'SHORT', f'BOOK_{int(PROFIT_BOOK_PCT*100)}PCT_REENTRY')
        else:
            msg = (
                f"⏸ Re-entry blocked by structural filter\n"
                f"Direction: {direction} | Spot: {close}\n"
                f"SPH: {SPH} | SPL: {SPL}\n"
                f"Staying flat — normal SPH/SPL breakout logic active."
            )
            logger.info(msg)
            self.send_telegram(msg)

    def _exit_position(self, t, close, pos, reason):
        """Exit the current position — closes all legs and logs the trade."""
        legs = pos.get('legs', [])
        total_pnl = 0
        legs_exit_info = []

        if legs:
            # Multi-leg exit (SYNTHETIC, ZEBRA, or single-leg with legs format)
            for leg in legs:
                symbol = leg['symbol']
                entry_price = leg.get('entry_price', 0)
                exit_ltp = self.get_option_ltp(symbol) if symbol else 0
                exit_side = 'SELL' if leg['side'] == 'BUY' else 'BUY'
                qty = leg['qty']

                if leg['side'] == 'BUY':
                    leg_pnl = (exit_ltp - entry_price) * qty
                else:
                    leg_pnl = (entry_price - exit_ltp) * qty

                total_pnl += leg_pnl

                if self.mode == 'LIVE':
                    try:
                        self.place_order(symbol, exit_side, qty)
                        logger.info(f"✅ LIVE Exit: {exit_side} {symbol} x {qty}")
                    except Exception as e:
                        logger.error(f"❌ Exit Order Failed: {e}")
                        self.send_telegram(f"❌ Exit Order Failed ({exit_side} {symbol}): {e}")

                legs_exit_info.append(f"  {exit_side} {symbol}: ₹{entry_price:.2f}→₹{exit_ltp:.2f} (₹{leg_pnl:+,.2f})")
        else:
            # Legacy single-leg fallback (old positions without 'legs' key)
            symbol = pos.get('symbol')
            entry_price = pos.get('entry_price', 0)
            exit_ltp = self.get_option_ltp(symbol) if symbol else 0
            pnl_points = exit_ltp - entry_price
            total_pnl = pnl_points * self.qty

            if self.mode == 'LIVE':
                try:
                    if symbol:
                        self.place_order(symbol, 'SELL', self.qty)
                    logger.info(f"✅ LIVE Exit: SELL {symbol} x {self.qty}")
                except Exception as e:
                    logger.error(f"❌ Exit Order Failed: {e}")
                    self.send_telegram(f"❌ Exit Order Failed: {e}")

            legs_exit_info.append(f"  SELL {symbol}: ₹{entry_price:.2f}→₹{exit_ltp:.2f}")

        # ── TELEGRAM: EXIT ──
        emoji = "🟢" if total_pnl >= 0 else "🔴"
        exit_detail = "\n".join(legs_exit_info)
        exit_msg = (
            f"🛑 **TRADE EXIT**\n"
            f"Direction: {pos['type']} ({pos.get('order_type', 'BUY')})\n"
            f"Reason: {reason}\n"
            f"{exit_detail}\n"
            f"{emoji} **Total PnL: ₹{total_pnl:,.2f}**\n"
            f"Time: {t.strftime('%d-%b %H:%M:%S')}"
        )
        logger.info(exit_msg)
        self.send_telegram(exit_msg)

        self.log_completed_trade({
            'Entry Time': str(pos.get('time', '')),
            'Exit Time': str(t),
            'Type': pos['type'],
            'Order Type': pos.get('order_type', 'BUY'),
            'Symbol': pos.get('symbol', ''),
            'Strike': pos.get('strike', ''),
            'Opt Type': pos.get('opt_type', ''),
            'Entry Price': pos.get('entry_price', 0),
            'Exit Price': exit_ltp if not legs else 0,
            'PnL Value': total_pnl,
            'Reason': reason
        })

        self.positions = []
        self._clear_gap_sl_lock()
        self.save_state()

    # ═══════════════════════════════════════════════
    # EXPIRY DAY HANDLING
    # ═══════════════════════════════════════════════

    def handle_expiry_exit(self, t, close):
        """
        On expiry day at 15:15:
        1. Force-close any open position.
        2. Refresh instruments to pick the new next-week expiry.
        3. Wait for the next SPH/SPL signal to re-enter.
        """
        if not self.positions:
            logger.info("📅 Expiry Day: No positions to exit.")
            return

        pos = self.positions[0]
        symbol = pos.get('symbol')
        entry_price = pos.get('entry_price', 0)
        exit_ltp = self.get_option_ltp(symbol) if symbol else 0

        pnl_points = exit_ltp - entry_price
        total_pnl = pnl_points * self.qty

        # Exit
        if self.mode == 'LIVE':
            try:
                if symbol:
                    self.place_order(symbol, 'SELL', self.qty)
                logger.info(f"✅ Expiry Exit: SELL {symbol}")
            except Exception as e:
                logger.error(f"❌ Expiry Exit Failed: {e}")
                self.send_telegram(f"❌ Expiry Exit Order Failed: {e}")

        emoji = "🟢" if total_pnl >= 0 else "🔴"
        exp_msg = (
            f"📅 **EXPIRY DAY EXIT**\n"
            f"Symbol: {symbol}\n"
            f"Entry: ₹{entry_price:.2f} → Exit: ₹{exit_ltp:.2f}\n"
            f"{emoji} **PnL: ₹{total_pnl:,.2f}**\n"
            f"Time: {t.strftime('%d-%b %H:%M:%S')}\n"
            f"⚠️ Will re-enter on next signal with new expiry."
        )
        logger.info(exp_msg)
        self.send_telegram(exp_msg)

        self.log_completed_trade({
            'Entry Time': str(pos.get('time', '')),
            'Exit Time': str(t),
            'Type': pos['type'],
            'Symbol': symbol,
            'Strike': pos.get('strike', ''),
            'Opt Type': pos.get('opt_type', ''),
            'Entry Price': entry_price,
            'Exit Price': exit_ltp,
            'PnL Points': pnl_points,
            'PnL Value': total_pnl,
            'Reason': 'EXPIRY_EXIT'
        })

        self.positions = []

        # Refresh instruments for new next-week expiry
        logger.info("🔄 Refreshing instruments for new expiry...")
        self.setup_instruments()
        self.save_state()

    def handle_expiry_rollover(self, t, close):
        """
        Pre-expiry rollover: Exit current position, refresh to new expiry,
        then immediately re-enter the SAME direction with new expiry options.
        Triggers on Friday (or earlier) when held options are ≤4 days from expiry.
        """
        if not self.positions:
            return

        pos = self.positions[0]
        old_direction = pos['type']
        symbol = pos.get('symbol')
        entry_price = pos.get('entry_price', 0)
        old_expiry = pos.get('expiry', '')
        exit_ltp = self.get_option_ltp(symbol) if symbol else 0

        pnl_points = exit_ltp - entry_price
        total_pnl = pnl_points * self.qty

        # Exit current position
        if self.mode == 'LIVE':
            try:
                if symbol:
                    self.place_order(symbol, 'SELL', self.qty)
                logger.info(f"✅ Rollover Exit: SELL {symbol}")
            except Exception as e:
                logger.error(f"❌ Rollover Exit Failed: {e}")
                self.send_telegram(f"❌ Rollover Exit Failed: {e}")
                return

        emoji = "🟢" if total_pnl >= 0 else "🔴"

        self.log_completed_trade({
            'Entry Time': str(pos.get('time', '')),
            'Exit Time': str(t),
            'Type': old_direction,
            'Symbol': symbol,
            'Strike': pos.get('strike', ''),
            'Opt Type': pos.get('opt_type', ''),
            'Entry Price': entry_price,
            'Exit Price': exit_ltp,
            'PnL Points': pnl_points,
            'PnL Value': total_pnl,
            'Reason': 'PRE_EXPIRY_ROLLOVER'
        })

        self.positions = []

        # Refresh instruments → picks new next-week expiry
        self.setup_instruments()

        # Re-enter same direction with new expiry
        reason = f'ROLLOVER_{old_expiry}_to_{self.expiry_date}'
        self._enter_position(t, close, old_direction, reason)

        roll_msg = (
            f"🔄 **PRE-EXPIRY ROLLOVER**\n"
            f"Old Expiry: {old_expiry} → New: {self.expiry_date}\n"
            f"Direction: {old_direction} (maintained)\n"
            f"Old: {symbol} @ ₹{entry_price:.2f} → ₹{exit_ltp:.2f}\n"
            f"{emoji} Closed PnL: ₹{total_pnl:,.2f}\n"
            f"New: {self.positions[0]['symbol'] if self.positions else 'FAILED'} @ ₹{self.positions[0].get('entry_price', 0):.2f if self.positions else 0}\n"
            f"Time: {t.strftime('%d-%b %H:%M:%S')}"
        )
        logger.info(roll_msg)
        self.send_telegram(roll_msg)

    # ═══════════════════════════════════════════════
    # ORDER PLACEMENT
    # ═══════════════════════════════════════════════

    def place_order(self, symbol, side, quantity):
        """Place a MARKET order via Kite. Only called when mode == 'LIVE'."""
        try:
            txn = (
                self.kite.TRANSACTION_TYPE_SELL if side == 'SELL'
                else self.kite.TRANSACTION_TYPE_BUY
            )
            order_id = self.kite.place_order(
                tradingsymbol=symbol,
                exchange=self.kite.EXCHANGE_NFO,
                transaction_type=txn,
                quantity=quantity,
                order_type=self.kite.ORDER_TYPE_MARKET,
                product=self.kite.PRODUCT_NRML,
                variety=self.kite.VARIETY_REGULAR,
                market_protection=-1,
            )
            logger.info(f"Order OK: {side} {symbol} x {quantity} | ID: {order_id}")
            return order_id
        except Exception as e:
            logger.error(f"Kite Order Error ({symbol}): {e}")
            raise e

    # ═══════════════════════════════════════════════
    # STATE MANAGEMENT (Crash Recovery)
    # ═══════════════════════════════════════════════

    def save_state(self):
        """Persist bot state to disk for crash recovery."""
        try:
            state = {
                'positions': self.positions,
                'sph': float(self.sph) if not pd.isna(self.sph) else None,
                'spl': float(self.spl) if not pd.isna(self.spl) else None,
                'expiry_date': str(self.expiry_date) if self.expiry_date else None,
                'last_exec_15min': self.last_exec_15min,
                'gap_sl_locked': self.gap_sl_locked,
                'gap_sl_locked_value': self.gap_sl_locked_value,
                'pending_live_direction': self.pending_live_direction,
            }
            with open('bot_v2_state.json', 'w') as f:
                json.dump(state, f, indent=2)
            logger.info("💾 State Saved.")
        except Exception as e:
            logger.error(f"Save State Failed: {e}")

    def load_state(self):
        """Load bot state from disk. Validates saved positions against current instruments."""
        try:
            if not os.path.exists('bot_v2_state.json'):
                return

            with open('bot_v2_state.json', 'r') as f:
                state = json.load(f)

            self.sph = state.get('sph', np.nan)
            self.spl = state.get('spl', np.nan)
            if self.sph is None:
                self.sph = np.nan
            if self.spl is None:
                self.spl = np.nan

            self.gap_sl_locked = state.get('gap_sl_locked')
            self.gap_sl_locked_value = state.get('gap_sl_locked_value')
            self.pending_live_direction = state.get('pending_live_direction')

            # Restore positions — validate symbol still exists
            loaded_positions = state.get('positions', [])
            valid_positions = []

            if not self.instruments_df.empty:
                valid_symbols = set(self.instruments_df['tradingsymbol'].values)
                for pos in loaded_positions:
                    sym = pos.get('symbol')
                    if sym and sym in valid_symbols:
                        valid_positions.append(pos)
                    else:
                        # Symbol might be from a different expiry — still check via LTP
                        try:
                            ltp = self.get_option_ltp(sym)
                            if ltp > 0:
                                valid_positions.append(pos)
                                logger.info(f"Position {sym} from different expiry, still valid (LTP={ltp})")
                            else:
                                logger.warning(f"Discarding expired position: {sym}")
                        except:
                            logger.warning(f"Discarding expired position: {sym}")

            self.positions = valid_positions

            if self.positions:
                pos = self.positions[0]
                logger.info(f"⚡ Resumed: {pos['type']} in {pos['symbol']} @ ₹{pos.get('entry_price', 0):.2f}")

        except Exception as e:
            logger.error(f"Load State Failed: {e}")

    def _saved_position_qty(self, pos) -> int | None:
        legs = pos.get('legs') or []
        if legs:
            return int(legs[0].get('qty', 0) or 0)
        return int(pos.get('qty', 0) or 0) or None

    def _resync_live_position_if_needed(self):
        """
        LIVE: if state was saved from paper/sim at a different lot size, do not
        resume those legs — place a fresh broker entry at self.qty instead.
        Retries via pending_live_direction if the broker order failed.
        """
        if self.mode != 'LIVE':
            return

        if not self.positions and self.pending_live_direction:
            direction = self.pending_live_direction
            ref_price = None
            if not self.df_15min.empty:
                try:
                    ref_price = float(self.df_15min['close'].iloc[-1])
                except Exception:
                    ref_price = None
            if ref_price and ref_price > 0:
                logger.info(f"Retrying pending LIVE fresh {direction} entry @ {self.qty} qty")
                self._enter_position(
                    datetime.datetime.now(),
                    ref_price,
                    direction,
                    'LIVE_FRESH_ENTRY_RETRY',
                )
                if self.positions:
                    self.pending_live_direction = None
                    self.save_state()
            return

        if not self.positions:
            return

        pos = self.positions[0]
        saved_qty = self._saved_position_qty(pos)
        if saved_qty is not None and saved_qty == self.qty:
            return

        direction = pos.get('type', 'SHORT')
        symbol = pos.get('symbol', '?')
        logger.warning(
            f"LIVE qty mismatch (saved {saved_qty} vs {self.qty}) — "
            f"discarding paper state for {symbol}, fresh {direction} entry"
        )
        self.send_telegram(
            f"🔄 **LIVE FRESH ENTRY**\n"
            f"Discarding paper position ({saved_qty} qty).\n"
            f"Placing new **{direction}** @ **{self.qty} qty** ({self.num_lots} lots)…"
        )

        self.pending_live_direction = direction
        self.positions = []
        self.save_state()

        ref_price = None
        if not self.df_15min.empty:
            try:
                ref_price = float(self.df_15min['close'].iloc[-1])
            except Exception:
                ref_price = None
        if not ref_price:
            ref_price = float(pos.get('spot_price', 0) or 0)
        if ref_price <= 0:
            logger.error("Cannot fresh-enter LIVE: no reference spot price")
            self.send_telegram("❌ LIVE fresh entry failed: no spot price")
            return

        self._enter_position(
            datetime.datetime.now(),
            ref_price,
            direction,
            'LIVE_FRESH_ENTRY',
        )
        if self.positions:
            self.pending_live_direction = None
            self.save_state()

    # ═══════════════════════════════════════════════
    # TELEGRAM INTEGRATION
    # ═══════════════════════════════════════════════
    #
    # Credentials: TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID (environment variables).
    #
    # ── OUTBOUND MESSAGES (Bot → You) ──
    #   🤖 Bot Started          — On startup
    #   🚀 New Entry            — On SPH/SPL breakout entry
    #   🛑 Trade Exit           — On reversal exit
    #   🔄 Trailing SL Updated  — When SPH/SPL pivot level changes
    #   ⏱  Status Update        — Every 15 minutes (if in a position)
    #   📅 Expiry Day Exit      — Forced exit on expiry day
    #   🚨 Websocket Down       — If no ticks for 60+ seconds
    #   ✅ Positions Closed     — On /ABORT command
    #
    # ── INBOUND COMMANDS (You → Bot) ──
    #   /STATUS  — Get current position & PnL
    #   /STOP    — Graceful shutdown (positions kept open)
    #   /ABORT   — Force-close all positions and shutdown
    # ═══════════════════════════════════════════════

    def send_telegram(self, message, retries=3, background=True):
        """Send a message to the configured Telegram chat (non-blocking by default)."""
        if not self.telegram_bot_token or not self.telegram_chat_id:
            return
        if "REPLACE" in self.telegram_bot_token:
            return

        def _deliver():
            url = f"https://api.telegram.org/bot{self.telegram_bot_token}/sendMessage"
            payload = {"chat_id": self.telegram_chat_id, "text": message}
            for attempt in range(1, retries + 1):
                try:
                    response = requests.post(url, data=payload, timeout=(5, 15))
                    if response.status_code == 200:
                        return
                    logger.error(
                        f"Telegram Send HTTP {response.status_code}: {response.text[:200]}"
                    )
                except Exception as e:
                    logger.error(f"Telegram Send Error (attempt {attempt}/{retries}): {e}")
                if attempt < retries:
                    time.sleep(min(2 ** attempt, 10))

        if background:
            threading.Thread(target=_deliver, daemon=True).start()
        else:
            _deliver()

    def send_sl_update(self, new_sl, pos_type):
        """Notify when the trailing stop level changes."""
        msg = (
            f"🔄 **TRAILING SL UPDATED**\n"
            f"Position: {pos_type}\n"
            f"New SL Level: {new_sl:.2f}\n"
            f"Time: {datetime.datetime.now().strftime('%H:%M:%S')}"
        )
        self.send_telegram(msg)

    def send_position_update(self):
        """Send current position status (heartbeat or /STATUS response)."""
        if not self.positions:
            self.send_telegram("📊 Status: FLAT — No Active Position")
            return

        pos = self.positions[0]
        symbol = pos.get('symbol')
        entry_price = pos.get('entry_price', 0)
        current_ltp = self.get_option_ltp(symbol) if symbol else 0

        pnl_points = current_ltp - entry_price
        total_pnl = pnl_points * self.qty
        emoji = "🟢" if total_pnl >= 0 else "🔴"

        sl = self.spl if pos['type'] == 'LONG' else self.sph

        msg = (
            f"⏱ **STATUS UPDATE**\n"
            f"Time: {datetime.datetime.now().strftime('%d-%b %H:%M:%S')}\n"
            f"Position: {pos['type']}\n"
            f"Symbol: {symbol}\n"
            f"Entry: ₹{entry_price:.2f} → LTP: ₹{current_ltp:.2f}\n"
            f"🛑 Trailing SL: {sl}\n"
            f"{emoji} Unrealized PnL: ₹{total_pnl:,.2f} ({pnl_points:.2f} pts)\n"
            f"Expiry: {pos.get('expiry', self.expiry_date)}"
        )
        self.send_telegram(msg)

    def start_telegram_listener(self):
        """Start background thread to listen for Telegram commands."""
        if not self.telegram_bot_token or not self.telegram_chat_id:
            logger.warning("Telegram listener skipped: set TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID")
            return
        logger.info("Starting Telegram Command Listener...")
        t = threading.Thread(target=self._telegram_listener_boot, daemon=True)
        t.start()

    def _telegram_listener_boot(self):
        """Flush stale commands then poll — kept off the main thread."""
        self._flush_telegram_updates()
        self._telegram_polling_loop()

    def _flush_telegram_updates(self):
        """Discard all pending Telegram messages so old /STOP commands are ignored."""
        try:
            url = f"https://api.telegram.org/bot{self.telegram_bot_token}/getUpdates"
            response = requests.get(url, params={'timeout': 1}, timeout=5)
            if response.status_code == 200:
                data = response.json()
                results = data.get('result', [])
                if results:
                    # Set offset to last update + 1 to mark all as read
                    self.last_update_id = results[-1]['update_id']
                    # Confirm flush with Telegram API
                    requests.get(url, params={'offset': self.last_update_id + 1, 'timeout': 1}, timeout=5)
                    logger.info(f"🧹 Flushed {len(results)} old Telegram message(s)")
                else:
                    logger.info("🧹 No pending Telegram messages to flush")
        except Exception as e:
            logger.error(f"Telegram flush error: {e}")

    def _telegram_polling_loop(self):
        """
        Poll Telegram for commands:
        /STOP     — Shutdown (keep positions)
        /ABORT    — Close positions & shutdown
        /STATUS   — Show current position & pivots
        /CONFIG   — Show current order config
        /BUY      — Switch to BUY mode (next trade)
        /SELL     — Switch to SELL mode (next trade)
        /SYNTHETIC — Switch to SYNTHETIC mode (next trade)
        /ZEBRA    — Switch to ZEBRA mode (next trade)
        /LOTS N   — Change lot size to N (next trade)
        /ITM N    — Change ITM offset to N strikes
        /OTM N    — Change OTM offset to N strikes
        """
        while True:
            try:
                if self.stop_signal_received:
                    break

                url = f"https://api.telegram.org/bot{self.telegram_bot_token}/getUpdates"
                params = {'offset': self.last_update_id + 1, 'timeout': 30}

                response = requests.get(url, params=params, timeout=35)
                if response.status_code == 200:
                    data = response.json()
                    for update in data.get('result', []):
                        self.last_update_id = update['update_id']

                        message = update.get('message', {})
                        text = message.get('text', '').strip()
                        text_upper = text.upper()
                        chat_id = str(message.get('chat', {}).get('id', ''))

                        if chat_id == self.telegram_chat_id:
                            if text_upper == '/STOP':
                                logger.warning("🛑 STOP COMMAND from Telegram")
                                self.send_telegram(
                                    "⚠️ **STOP COMMAND RECEIVED.**\n"
                                    "Bot shutting down.\n"
                                    "Positions are kept open (not closed)."
                                )
                                self.stop_signal_received = True
                                os._exit(0)

                            elif text_upper == '/ABORT':
                                logger.warning("🛑 ABORT COMMAND from Telegram")
                                self.send_telegram(
                                    "⚠️ **ABORT COMMAND RECEIVED.**\n"
                                    "Closing all positions & stopping."
                                )
                                self.close_all_positions()
                                self.stop_signal_received = True
                                os._exit(0)

                            elif text_upper == '/STATUS':
                                self.send_position_update()

                            elif text_upper == '/CONFIG':
                                self._send_config()

                            elif text_upper in ('/BUY', '/SELL', '/SYNTHETIC', '/ZEBRA'):
                                new_type = text_upper[1:]  # Remove the /
                                old_type = self.order_type
                                self.order_type = new_type
                                logger.info(f"⚙️ Order type changed: {old_type} → {new_type}")
                                self.send_telegram(
                                    f"⚙️ **ORDER TYPE CHANGED**\n"
                                    f"{old_type} → **{new_type}**\n"
                                    f"Applies to next trade entry."
                                )

                            elif text_upper.startswith('/LOTS'):
                                parts = text_upper.split()
                                if len(parts) == 2 and parts[1].isdigit():
                                    new_lots = int(parts[1])
                                    old_lots = self.num_lots
                                    self.num_lots = new_lots
                                    self.qty = self.lot_size * new_lots
                                    logger.info(f"⚙️ Lots changed: {old_lots} → {new_lots} ({self.qty} qty)")
                                    self.send_telegram(
                                        f"⚙️ **LOTS CHANGED**\n"
                                        f"{old_lots} → **{new_lots}** lots ({self.qty} qty)\n"
                                        f"Applies to next trade entry."
                                    )
                                else:
                                    self.send_telegram("Usage: /LOTS 5")

                            elif text_upper.startswith('/ITM'):
                                parts = text_upper.split()
                                if len(parts) == 2 and parts[1].isdigit():
                                    new_itm = int(parts[1])
                                    self.itm_offset = new_itm
                                    logger.info(f"⚙️ ITM offset → {new_itm} strikes")
                                    self.send_telegram(f"⚙️ **ITM OFFSET → {new_itm}** strikes\nApplies to next trade.")
                                else:
                                    self.send_telegram("Usage: /ITM 2")

                            elif text_upper.startswith('/OTM'):
                                parts = text_upper.split()
                                if len(parts) == 2 and parts[1].isdigit():
                                    new_otm = int(parts[1])
                                    self.otm_offset = new_otm
                                    logger.info(f"⚙️ OTM offset → {new_otm} strikes")
                                    self.send_telegram(f"⚙️ **OTM OFFSET → {new_otm}** strikes\nApplies to next trade.")
                                else:
                                    self.send_telegram("Usage: /OTM 2")

                time.sleep(1)
            except Exception as e:
                logger.error(f"Telegram Listener Error: {e}")
                time.sleep(5)

    def _send_config(self):
        """Send current order configuration via Telegram."""
        pos_info = "FLAT"
        if self.positions:
            pos = self.positions[0]
            pos_info = f"{pos['type']} ({pos.get('order_type', 'BUY')})"

        self.send_telegram(
            f"⚙️ **BOT CONFIG**\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"Mode: {self.mode}\n"
            f"Order Type: **{self.order_type}**\n"
            f"Lots: **{self.num_lots}** ({self.qty} qty)\n"
            f"ITM Offset: {self.itm_offset} strikes\n"
            f"OTM Offset: {self.otm_offset} strikes\n"
            f"ZEBRA Deep ITM: {self.zebra_deep_itm} strikes\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"Position: {pos_info}\n"
            f"SPH: {self.sph} | SPL: {self.spl}\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"Commands:\n"
            f"/BUY /SELL /SYNTHETIC /ZEBRA\n"
            f"/LOTS N  /ITM N  /OTM N"
        )

    def close_all_positions(self):
        """Force-close all open positions (used by /ABORT command)."""
        if not self.positions:
            return
        try:
            pos = self.positions[0]
            symbol = pos.get('symbol')

            if self.mode == 'LIVE' and symbol:
                self.place_order(symbol, 'SELL', self.qty)
                logger.info("✅ Force Closed All Positions.")
                self.send_telegram("✅ All Positions Force Closed.")

            self.positions = []
            self.save_state()
        except Exception as e:
            logger.error(f"Force Close Failed: {e}")
            self.send_telegram(f"❌ Force Close Failed: {e}")

    # ═══════════════════════════════════════════════
    # WATCHDOG & HEARTBEAT
    # ═══════════════════════════════════════════════

    def check_heartbeat(self):
        """Background loop: check ticker health & send periodic updates."""
        now = datetime.datetime.now()

        # --- EOD disconnect fallback (if no tick arrived exactly at 15:30) ---
        disconnect_h, disconnect_m = TICKER_DISCONNECT_TIME
        if (
            now.hour == disconnect_h
            and now.minute >= disconnect_m
            and self.ticker_connected
            and self._eod_disconnect_date != now.date()
        ):
            self._handle_eod_disconnect(now)

        # --- Morning connect: Mon–Fri from 8:00 AM ---
        self._maybe_connect_morning_session(now)

        # --- Daily ticker reconnect after morning login (no second login) ---
        if not hasattr(self, "_session_refresh_date"):
            self._session_refresh_date = None
        if (
            self._session_refresh_date == now.date()
            and not self.ticker_connected
            and self.is_ticker_session_hours(now)
            and 8 <= now.hour < 10
            and self.is_trading_day(now)
        ):
            token = read_access_token()
            if token and token_is_valid(self.api_key, token):
                self.connect_ticker("post-login connect")

        # Watchdog: only during active websocket session (not after hours / weekends)
        if self.ticker_connected and self.is_ticker_session_hours(now):
            stale_seconds = (now - self.last_tick_time).total_seconds()
            if stale_seconds > 60:
                logger.warning(f"⚠️ WATCHDOG: No ticks for {int(stale_seconds)}s")
                self.send_telegram(
                    f"🚨 **WEBSOCKET DOWN**\n"
                    f"No data for {int(stale_seconds)}s!\n"
                    f"Last tick: {self.last_tick_time.strftime('%H:%M:%S')}\n"
                    f"Attempting auto-reconnect..."
                )
                try:
                    self.disconnect_ticker("watchdog reconnect")
                    self.connect_ticker("watchdog reconnect")
                except Exception as e:
                    logger.error(f"Reconnect Failed: {e}")
                    self.send_telegram(f"❌ Reconnect FAILED: {e}\nManual restart needed!")

        # 15-min heartbeat (only while websocket session is active)
        if (
            self.ticker_connected
            and self.is_ticker_session_hours(now)
            and (now - self.last_heartbeat_time).total_seconds() >= 900
        ):
            self.send_position_update()
            self.last_heartbeat_time = now

    # ═══════════════════════════════════════════════
    # TRADE LOGGING (Excel)
    # ═══════════════════════════════════════════════

    def log_completed_trade(self, trade_dict):
        """Append completed trade to Excel log with color coding."""
        filename = "Completed_Trades_V2.xlsx"

        if os.path.exists(filename):
            df_existing = pd.read_excel(filename)
        else:
            df_existing = pd.DataFrame()

        df_new = pd.DataFrame([trade_dict])
        df_final = pd.concat([df_existing, df_new], ignore_index=True)

        if 'PnL Value' in df_final.columns:
            df_final['PnL Value'] = pd.to_numeric(df_final['PnL Value'], errors='coerce')
            df_final['Cumulative PnL'] = df_final['PnL Value'].cumsum()

        df_final.to_excel(filename, index=False)
        logger.info(f"📝 Trade logged to {filename}")

        # Color-code PnL column
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font

            wb = load_workbook(filename)
            ws = wb.active

            pnl_col = None
            for cell in ws[1]:
                if cell.value == 'PnL Value':
                    pnl_col = cell.column

            if pnl_col:
                green = Font(color="008000", bold=True)
                red = Font(color="FF0000", bold=True)
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=pnl_col)
                    try:
                        if float(cell.value) >= 0:
                            cell.font = green
                        else:
                            cell.font = red
                    except:
                        pass

            wb.save(filename)
        except:
            pass

    # ═══════════════════════════════════════════════
    # DATA INITIALIZATION (Warmup)
    # ═══════════════════════════════════════════════

    def initialize_data(self):
        """Fetch native 15-min candles from Kite for pivot warmup."""
        now = datetime.datetime.now()
        logger.info("Initializing Data (Historical Fetch)...")

        to_date = now
        from_date = to_date - datetime.timedelta(days=10)

        try:
            # Fetch native 15-min candles directly (matches TradingView exactly)
            records = self.kite.historical_data(
                self.instrument_token, from_date, to_date, "15minute"
            )
            if not records:
                return

            df_hist = pd.DataFrame(records)
            df_hist['date'] = pd.to_datetime(df_hist['date']).dt.tz_localize(None)
            df_hist.set_index('date', inplace=True)

            self.df_15min = df_hist
            self.recalculate_pivots()
            logger.info(f"✅ Loaded {len(self.df_15min)} native 15-min candles for warmup")

        except Exception as e:
            logger.error(f"Initialization Error: {e}")

    # ═══════════════════════════════════════════════
    # MAIN ENTRY POINT
    # ═══════════════════════════════════════════════

    def start(self):
        """Boot the bot — connect, warmup, and run forever."""
        logger.info(f"🤖 Starting SPH/SPL Pivot Bot | Mode: {self.mode}")

        # ── TELEGRAM: STARTUP ──
        exec_mode = (
            "15-min candle close"
            if EXECUTE_ON_15MIN_CLOSE
            else "tick-by-tick (legacy)"
        )
        startup_msg = (
            f"🤖 **BOT V2 STARTED**\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"Mode: {self.mode}\n"
            f"Strategy: SPH/SPL Breakout (15-Min)\n"
            f"Signals: {exec_mode}\n"
            f"Reversal cooldown: {REVERSAL_COOLDOWN_SEC}s\n"
            f"Session SL cap: {'ON' if SESSION_SL_CAP_ENABLED else 'OFF'}\n"
            f"Style: Positional (Always-In Reversal)\n"
            f"Orders: BUY Only (1 ITM)\n"
            f"Qty: {self.qty} ({self.qty // self.lot_size} lots)\n"
            f"Expiry: {self.expiry_date}\n"
            f"━━━━━━━━━━━━━━━━━━━\n"
            f"Commands: /STATUS /STOP /ABORT\n"
            f"Status: Monitoring 🟢"
        )
        # Fire-and-forget with retries — must not block ticker connect on network hangs.
        self.send_telegram(startup_msg, retries=5)

        # Start Telegram listener early so /STATUS works while data warms up.
        self.start_telegram_listener()

        # Start Ticker (only during Mon–Fri 8:00–15:30 IST)
        if self.is_ticker_session_hours():
            self.connect_ticker("startup")
        else:
            logger.info(
                "Outside websocket hours (Mon–Fri 8:00–15:30 IST) — "
                "ticker idle until morning connect"
            )

        # Initialize data (pivot warmup)
        self.initialize_data()

        # LIVE: replace paper/sim position saved at wrong lot size with fresh broker entry
        self._resync_live_position_if_needed()

        # Optional: force startup direction instead of blindly resuming old state.
        if STARTUP_DIRECTION_OVERRIDE in ("LONG", "SHORT"):
            desired = STARTUP_DIRECTION_OVERRIDE
            if self.positions:
                current = self.positions[0].get('type')
                if current != desired:
                    ref_price = None
                    if not self.df_15min.empty:
                        try:
                            ref_price = float(self.df_15min['close'].iloc[-1])
                        except Exception:
                            ref_price = None
                    if ref_price is None:
                        ref_price = float(self.positions[0].get('spot_price', 0) or 0)
                    if ref_price > 0:
                        logger.info(
                            f"🎯 STARTUP OVERRIDE: {current} → {desired} at approx spot {ref_price}"
                        )
                        self._exit_position(
                            datetime.datetime.now(),
                            ref_price,
                            self.positions[0],
                            f"STARTUP_OVERRIDE_{desired}",
                        )
                        if not self.positions:
                            self._enter_position(
                                datetime.datetime.now(),
                                ref_price,
                                desired,
                                f"STARTUP_OVERRIDE_{desired}",
                            )
                    else:
                        logger.warning(
                            "⚠️ STARTUP OVERRIDE requested but no valid reference price found."
                        )

        if self.positions:
            logger.info("⚡ Resuming active position...")
            if SESSION_SL_CAP_ENABLED:
                self._apply_session_sl_cap()
            self.send_position_update()

        # Second startup ping after init (covers first-send network blips).
        threading.Timer(
            10.0,
            lambda: self.send_telegram(
                f"🟢 **BOT ONLINE**\n"
                f"Mode: {self.mode} | Expiry: {self.expiry_date}\n"
                f"Position: "
                f"{self.positions[0]['type'] + ' ' + self.positions[0]['symbol'] if self.positions else 'FLAT'}\n"
                f"SPH: {self.sph} | SPL: {self.spl}",
                retries=3,
            ),
        ).start()

        # Keep-alive loop
        logger.info("Bot running. Entering keep-alive loop.")
        while True:
            self.check_heartbeat()
            time.sleep(1)


# ═══════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════

if __name__ == '__main__':
    bot = StructuralPivotBot(mode=BOT_MODE)
    bot.start()
